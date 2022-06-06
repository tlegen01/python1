import openpyxl
from bottle import get
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

kniga1 = Workbook()
stranica1 = kniga1.active

var_range = range(1, 26)
stroka_1 = []
stroka1 = ""
for i in var_range:
    stroka1 = f"_{i}"
    stroka_1.append(stroka1)

stroka_2 = []
stroka2 = ""
for j in var_range:
    stroka2 = f"A_{j}"
    stroka_2.append(stroka2)

stroka_final = []
stroka_final.append(stroka_1)
stroka_final.append(stroka_2)
# print(stroke_final)

for row in range(0, len(stroka_final)):
    for col in range(0, len(stroka_final[row])):
        # print(stroka_final[row])
        # col_letter = get_column_letter(col)
        stranica1[f"{get_column_letter(col + 1)}{row + 1}"] = stroka_final[row][col]

for col in range(0, len(stroka_final)):
    for row in range(0, len(stroka_final[col])):
        stranica1.cell(row=row + 4, column=col + 1).value = stroka_final[col][row]
print(stroka_final)

kniga1.save("home_work4.xlsx")