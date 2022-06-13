import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import xlsxwriter

dir_name = "/home/tlegen/Документы/GitHub/python1/home_work"
files = []
for filename in os.listdir(dir_name):
    files.append(filename)
data = []
file_name = []
for file in files:
    file_name1 = dir_name + "/" + file
    # file_name2.append(file_name1)
    # print(file_name2)
    # print(type(file_name1))
    worbook = openpyxl.load_workbook(file_name1)
    worksheet = worbook.active
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    # start = str(worksheet["C3"].value)
    # print(start)
    #
    external_array = []
    for row in range(1, max_row + 1):
        internal_array = []
        for col in range(1, max_column + 1):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ""
            internal_array.append(value)
            # if len(internal_array) < 1:
            #     continue
        external_array.append(internal_array)
    # print(external_array)
    data.append(external_array)
# print(data)

arr1 = [x[0] for x in data[0]]
# print(arr1)
arr2 = [x[2] for x in data[1]]
# print(arr2)
arr3 = [x[1] for x in data[2]]
# print(arr3)

len_max = len(arr1)
if len(arr2) > len_max:
    len_max = len(arr2)
if len(arr3) > len_max:
    len_max = len(arr3)
# print(len_max)

sheet1 = []
sheet2 = []
sheet3 = []
for i in range(0, len_max):
    try:
        elem1 = arr1[i]
    except Exception as error:
        elem1 = ''
    try:
        elem2 = arr2[i]
    except Exception as error:
        elem2 = ''
    try:
        elem3 = arr3[i]
    except Exception as error:
        elem3 = ''
    sheet1.append(elem1)
    sheet2.append(elem2)
    sheet3.append(elem3)
# print(sheet1)
# print(sheet2)
# print(sheet3)

book = openpyxl.Workbook()
book.remove(book.active)
sheet_1 = book.create_sheet("Stolbec1")
sheet_2 = book.create_sheet("Stolbec2")
sheet_3 = book.create_sheet("Stolbec3")
max_row1 = sheet_1.max_row
max_row2 = sheet_2.max_row
max_row3 = sheet_3.max_row
max_column1 = sheet_1.max_column
max_column2 = sheet_2.max_column
max_column3 = sheet_3.max_column
print(sheet2)

for row in range(0, len(sheet1)):
    col = 1
    sheet_1.cell(row=row + 1, column=col).value = sheet1[row]
for row in range(0, len(sheet2)):
    col = 1
    sheet_2.cell(row=row + 1, column=col).value = sheet2[row]
for row in range(0, len(sheet3)):
    col = 1
    sheet_3.cell(row=row + 1, column=col).value = sheet3[row]
book.save("hw_final2.xlsx")